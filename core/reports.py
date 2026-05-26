from __future__ import annotations
from io import BytesIO
from pathlib import Path
from datetime import datetime, date, time
import json, math
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEV = 'Ing. Christian Pocol, Ingeniero Electrónico'
BLUE = '172B85'
GREEN = '00A83B'
DARK = '263238'

def _is_missing(value) -> bool:
    if value is None:
        return True
    try:
        if value is pd.NaT or pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, float) and math.isnan(value):
        return True
    return False

def _safe_cell(value):
    if _is_missing(value):
        return ''
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ''
        try:
            if value.tzinfo is not None:
                value = value.tz_convert(None)
        except Exception:
            try: value = value.tz_localize(None)
            except Exception: pass
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, datetime):
        if value.tzinfo is not None: value = value.replace(tzinfo=None)
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date): return value.strftime('%Y-%m-%d')
    if isinstance(value, time): return value.replace(tzinfo=None).strftime('%H:%M:%S') if value.tzinfo else value.strftime('%H:%M:%S')
    if isinstance(value, (dict, list, tuple, set)):
        try: return json.dumps(value, ensure_ascii=False, default=str)
        except Exception: return str(value)
    try:
        if hasattr(value, 'item'): value = value.item()
    except Exception: pass
    if isinstance(value, str): return value.replace('T', ' ').replace('Z', '')
    return value

def _write_df(ws, df: pd.DataFrame | None):
    if df is None or df.empty:
        ws.append(['Sin datos disponibles'])
        return
    clean = df.copy()
    clean.columns = [str(c) for c in clean.columns]
    ws.append(list(clean.columns))
    for row in clean.itertuples(index=False, name=None):
        ws.append([_safe_cell(v) for v in row])

def _format_ws(ws):
    ws.freeze_panes = 'A2'
    header_fill = PatternFill('solid', fgColor=BLUE)
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        values = [str(c.value) if c.value is not None else '' for c in column_cells]
        max_len = min(max([len(v) for v in values] + [10]) + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len

def excel_bytes(summary: pd.DataFrame, submissions: pd.DataFrame, history: pd.DataFrame | None = None, followups: pd.DataFrame | None = None, modules: pd.DataFrame | None = None, module_items: pd.DataFrame | None = None, module_matrix: pd.DataFrame | None = None) -> bytes:
    bio = BytesIO()
    wb = Workbook()
    sheets = [
        ('Resumen estudiantes', summary),
        ('Entregas detalle', submissions),
        ('Historial riesgo', history),
        ('Bitacora seguimiento', followups),
        ('Modulos', modules),
        ('Items por modulo', module_items),
        ('Matriz modulos', module_matrix),
    ]
    ws = wb.active
    ws.title = sheets[0][0]
    _write_df(ws, sheets[0][1])
    for title, df in sheets[1:]:
        wsx = wb.create_sheet(title[:31])
        _write_df(wsx, df)
    if summary is not None and not summary.empty:
        ws_kpi = wb.create_sheet('Indicadores ejecutivos')
        counts = summary['riesgo_integral'].value_counts().to_dict() if 'riesgo_integral' in summary else {}
        kpis = [
            ['Indicador','Valor'],
            ['Total estudiantes', len(summary)],
            ['Riesgo alto', counts.get('Alto',0)],
            ['Riesgo medio', counts.get('Medio',0)],
            ['Riesgo bajo', counts.get('Bajo',0)],
            ['Avance promedio', round(float(summary.get('porcentaje_avance', pd.Series([0])).mean()),2)],
            ['Horas promedio', round(float(summary.get('tiempo_total_horas', pd.Series([0])).mean()),2)],
            ['Pendientes totales', int(summary.get('pendientes', pd.Series([0])).sum())],
            ['Atrasadas totales', int(summary.get('atrasadas', pd.Series([0])).sum())],
        ]
        for r in kpis: ws_kpi.append(r)
    for wsx in wb.worksheets:
        _format_ws(wsx)
    wb.save(bio)
    return bio.getvalue()

def _watermark(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica-Bold', 42)
    canvas.setFillColor(colors.Color(0.85, 0.85, 0.85, alpha=0.23))
    canvas.translate(5.5 * inch, 4.2 * inch)
    canvas.rotate(35)
    canvas.drawCentredString(0, 0, 'AVE - UVG')
    canvas.restoreState()
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(0.35 * inch, 0.25 * inch, f'Desarrollador: {DEV}')
    canvas.drawRightString(10.65 * inch, 0.25 * inch, f'Página {doc.page}')
    canvas.restoreState()

def _styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SmallAVE', fontSize=8, leading=10))
    styles.add(ParagraphStyle(name='TinyAVE', fontSize=6.5, leading=8))
    styles.add(ParagraphStyle(name='TitleAVE', fontSize=18, leading=22, alignment=1, textColor=colors.HexColor('#172B85')))
    styles.add(ParagraphStyle(name='SectionAVE', fontSize=12, leading=14, textColor=colors.HexColor('#172B85'), spaceBefore=10, spaceAfter=6))
    return styles

def _header(story, styles, title, course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg):
    header_data = []
    header_data.append(Image(logo_ave, width=1.25*inch, height=0.62*inch) if Path(logo_ave).exists() else '')
    header_data.append(Paragraph(title, styles['TitleAVE']))
    header_data.append(Image(logo_uvg, width=1.1*inch, height=0.62*inch) if Path(logo_uvg).exists() else '')
    ht = Table([header_data], colWidths=[1.4*inch, 7.2*inch, 1.4*inch])
    ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story += [ht, Spacer(1, 10)]
    story.append(Paragraph(f'<b>Curso:</b> {course_name}<br/><b>Sección:</b> {section_name}<br/><b>Fecha de análisis:</b> {analysis_date}<br/><b>Generado por:</b> {generated_by or "No especificado"}<br/><b>Desarrollador:</b> {DEV}', styles['SmallAVE']))
    story.append(Spacer(1, 10))

def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = _styles()
    story = []
    _header(story, styles, 'Informe Ejecutivo de Seguimiento Académico AVE', course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg)
    total = len(summary) if summary is not None else 0
    counts = summary['riesgo_integral'].value_counts().to_dict() if total and 'riesgo_integral' in summary.columns else {}
    kpi = [['Total estudiantes','Riesgo bajo','Riesgo medio','Riesgo alto','Prom. avance','Pendientes','Atrasadas'],[total, counts.get('Bajo',0), counts.get('Medio',0), counts.get('Alto',0), f"{summary['porcentaje_avance'].mean():.1f}%" if total and 'porcentaje_avance' in summary else '0%', int(summary.get('pendientes', pd.Series([0])).sum()) if total else 0, int(summary.get('atrasadas', pd.Series([0])).sum()) if total else 0]]
    kt = Table(kpi, colWidths=[1.25*inch]*7)
    kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('ALIGN',(0,0),(-1,-1),'CENTER'),('FONTSIZE',(0,0),(-1,-1),8)]))
    story += [kt, Spacer(1, 10)]
    story.append(Paragraph('<b>Interpretación ejecutiva:</b> el riesgo integral prioriza desconexión, déficit de horas, entregas pendientes, entregas atrasadas, bajo avance y reincidencia operativa. El listado siguiente incluye todos los casos de riesgo Alto y Medio para intervención del asesor.', styles['SmallAVE']))
    story.append(Spacer(1, 8))
    if summary is None or summary.empty:
        story.append(Paragraph('No hay estudiantes para mostrar.', styles['SmallAVE']))
    else:
        cols = ['estudiante','correo','horas_sin_actividad','riesgo_desconexion','pendientes','atrasadas','porcentaje_avance','puntaje_riesgo','riesgo_integral','segmento_ave','accion_recomendada']
        view = summary.copy()
        for c in cols:
            if c not in view.columns: view[c] = ''
        order = {'Alto':0,'Medio':1,'Bajo':2}
        view['_ord'] = view['riesgo_integral'].map(order).fillna(3)
        view = view[view['riesgo_integral'].isin(['Alto','Medio'])].sort_values(['_ord','puntaje_riesgo','horas_sin_actividad'], ascending=[True,False,False])
        if view.empty:
            view = summary.sort_values(['puntaje_riesgo','horas_sin_actividad'], ascending=[False,False])
        story.append(Paragraph(f'<b>Casos listados:</b> {len(view)} estudiantes. Riesgo alto: {counts.get("Alto",0)}. Riesgo medio: {counts.get("Medio",0)}.', styles['SmallAVE']))
        data = [['Estudiante','Correo','Hrs sin act.','Riesgo conexión','Pend.','Atr.','Avance %','Puntaje','Riesgo','Segmento','Acción']] + view[cols].fillna('').astype(str).values.tolist()
        table = Table(data, repeatRows=1, colWidths=[1.4*inch,1.45*inch,0.62*inch,0.85*inch,0.4*inch,0.4*inch,0.55*inch,0.52*inch,0.55*inch,1.0*inch,1.35*inch])
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6),('VALIGN',(0,0),(-1,-1),'TOP'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F7F9FB')])]))
        story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph('Nota: los datos dependen de los permisos del token y de los registros disponibles en Canvas. Se recomienda complementar la lectura con la bitácora de seguimiento del asesor.', styles['SmallAVE']))
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()

def individual_pdf_bytes(student_row: dict, submissions: pd.DataFrame, followups: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=30, bottomMargin=30)
    styles = _styles()
    story = []
    _header(story, styles, 'Ficha Individual de Seguimiento AVE', course_name, section_name, generated_by, analysis_date, logo_ave, logo_uvg)
    name = student_row.get('estudiante') or student_row.get('nombre') or 'Estudiante'
    story.append(Paragraph(f'<b>Estudiante:</b> {name}<br/><b>Correo:</b> {student_row.get("correo", "")}<br/><b>Riesgo integral:</b> {student_row.get("riesgo_integral", "")}<br/><b>Segmento AVE:</b> {student_row.get("segmento_ave", "")}<br/><b>Acción recomendada:</b> {student_row.get("accion_recomendada", "")}', styles['SmallAVE']))
    story.append(Spacer(1, 10))
    kpi = [['Horas sin actividad','Horas acumuladas','Horas esperadas','Déficit','Pendientes','Atrasadas','Avance %','Puntaje'],[student_row.get('horas_sin_actividad',''), student_row.get('tiempo_total_horas',''), student_row.get('horas_esperadas',''), student_row.get('deficit_horas',''), student_row.get('pendientes',''), student_row.get('atrasadas',''), student_row.get('porcentaje_avance',''), student_row.get('puntaje_riesgo','')]]
    kt = Table(kpi, colWidths=[0.9*inch]*8)
    kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172B85')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(0,0),(-1,-1),'CENTER')]))
    story += [kt, Spacer(1, 12)]
    story.append(Paragraph('Entregas del estudiante', styles['SectionAVE']))
    uid = student_row.get('user_id')
    sub = submissions[submissions['user_id'].astype(str)==str(uid)].copy() if submissions is not None and not submissions.empty and uid is not None else pd.DataFrame()
    if sub.empty:
        story.append(Paragraph('Sin detalle de entregas disponible para este estudiante.', styles['SmallAVE']))
    else:
        cols = [c for c in ['actividad','fecha_entrega','submitted_at','workflow_state','missing','late','score','puntos'] if c in sub.columns]
        data = [cols] + sub[cols].fillna('').astype(str).head(80).values.tolist()
        t = Table(data, repeatRows=1, colWidths=[2.1*inch,0.9*inch,0.9*inch,0.8*inch,0.45*inch,0.45*inch,0.45*inch,0.45*inch][:len(cols)])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph('Bitácora de seguimiento', styles['SectionAVE']))
    bit = followups[followups['user_id'].astype(str)==str(uid)].copy() if followups is not None and not followups.empty and uid is not None else pd.DataFrame()
    if bit.empty:
        story.append(Paragraph('Sin registros de seguimiento en la bitácora local.', styles['SmallAVE']))
    else:
        cols = [c for c in ['created_at','medio','motivo','resultado','proxima_accion','fecha_proxima_accion','observaciones','registrado_por'] if c in bit.columns]
        data = [cols] + bit[cols].fillna('').astype(str).head(30).values.tolist()
        t = Table(data, repeatRows=1, colWidths=[0.85*inch,0.7*inch,0.85*inch,0.9*inch,0.9*inch,0.75*inch,1.4*inch,0.8*inch][:len(cols)])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c757d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.22,colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t)
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()
